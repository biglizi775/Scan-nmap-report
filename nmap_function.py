import os
import re
import time

import openpyxl


def Volume_production(input_ipfile):
    book = openpyxl.Workbook()
    del book['Sheet']
    sheet = book.create_sheet('基本信息')
    sheet.cell(1, 2, 'PORT')
    sheet.cell(1, 3, 'STATE')
    sheet.cell(1, 4, 'SERVICE')

    ip_address = []
    with open(input_ipfile) as f1:
        for iprun in f1.readlines():
            if iprun != None:
                # 从文件中读取行数据时，会带换行符，使用strip函数去掉 换行符后存入列表
                ip_address.append(iprun.strip("\n"))
    # print(ip_address)
    f1.close()

    for ips in range(0, len(ip_address)):
        cols = 2
        # print(ips)
        # print(ip_address[ips])
        cmdip = str(ip_address[ips])
        # -sS 70-48581 -Pn -n --min-hostgroup 4 --min-parallelism 1024 --host-timeout 30
        # cmd = 'nmap ' + str(cmdip) + '  -p 70-48581 -Pn -n --min-hostgroup 4 --min-parallelism 1024 --host-timeout 30'
        cmd = 'nmap ' + str(cmdip) + '  -sS  -Pn -n --min-hostgroup 4 --min-parallelism 1024 --host-timeout 30'
        # print(cmd)
        res = os.popen(cmd)
        output_str = res.read()  # 获得输出字符串
        # print(type(output_str))
        # print(len(output_str))
        # print(output_str)
        output_str_dispose = "".join([s for s in output_str.splitlines(True) if s.strip()])  # 去掉空行
        # print(output_str_dispose)
        a = output_str_dispose.split("\n")  # 字符串按行输出
        # print(a)
        lines = 1

        iplocation = ''
        for i in range(0, len(a)):
            # print(a[i])
            if 'cp unknown' in a[i]:
                continue
            elif 'Nmap scan report for' in a[i]:
                # print(a[i] + ":" + str(len(a[i])) + ":" + str(i))
                result = re.findall(
                    r"\b(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\b",
                    a[i])
                # 提取ip
                iplocation = str(result[0])
                # print(str(result[0]))
            elif 'SERVICE' in a[i]:
                A1 = re.split(r" +", a[i])  # 按照多个空格分割
                # print(str(A1[0]), str(A1[1]), str(A1[2]))
                # print(str(a[i+1]))
                endbyte = ''
                nextline = 1
                while ('MAC Address:' not in endbyte) or ('Nmap done:' not in endbyte):
                    #print('I AM')
                    thisline = str(a[i + nextline])
                    endbyte = thisline
                    if ('MAC Address:' in endbyte) or ('Nmap done:' in endbyte):
                        break
                    # print(endbyte)

                    B2 = re.split(r" +", thisline)
                    # print(str(B2[0]), str(B2[1]), str(B2[2]), str(i))
                    if 'unknown' not in str(B2[1]):
                        sheet.cell(cols, 1, iplocation)
                        sheet.cell(cols, 2, str(B2[0]))
                        sheet.cell(cols, 3, str(B2[1]))#state
                        sheet.cell(cols, 4, str(B2[2]))
                        cols = cols + 1
                    nextline = nextline + 1

    book.save('C:\\Users\\31216\\Desktop\\test.xlsx')


# def main():
#     Volume_production('192.168.12.1-140')
if __name__ == '__main__':
    localtime = time.asctime(time.localtime(time.time()))
    print('1'+ str(localtime))
    Volume_production('C:\\Users\\31216\\Desktop\\testip.txt')
    print('2' + str(localtime))
